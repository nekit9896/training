"""
Общая логика разбора xlsx-отчётов с таблицей участков и длительностями режимов.

Используется отчётами о режиме СОУ и о режиме работы МТ.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, time, timedelta
from typing import Dict, List, Optional, Protocol, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from constants.test_constants import BaseTN3Constants as TestConst
from utils.helpers.report_xlsx_utils import (
    ReportTitleInfo,
    _stringify_cell,
    build_column_cells,
    get_report_column_headers,
    parse_report_title,
)

_DURATION_PARTS_H_MM_SS = 3
_DURATION_PARTS_MM_SS = 2


class ModeDurationReportConstants(Protocol):
    """Структурные константы xlsx-отчёта с таблицей участков и длительностями режимов."""

    REPORT_TITLE_ROW: int
    REPORT_COLUMN_HEADERS_ROW: int
    REPORT_DATA_FIRST_ROW: int
    TOTAL_WORK_DURATION_LABEL: str
    COL_SECTION: str
    REPORT_HEADER_PERIOD_PATTERN: str
    MODE_DURATION_COLUMNS: list


@dataclass(frozen=True)
class ModeDurationReportLayout:
    """Параметры разбора листа отчёта с длительностями режимов по участкам."""

    report_title_row: int
    report_column_headers_row: int
    report_data_first_row: int
    total_work_duration_label: str
    col_section: str
    header_period_pattern: str
    mode_duration_columns: Tuple[str, ...]

    @classmethod
    def from_constants(cls, constants: ModeDurationReportConstants) -> ModeDurationReportLayout:
        return cls(
            report_title_row=constants.REPORT_TITLE_ROW,
            report_column_headers_row=constants.REPORT_COLUMN_HEADERS_ROW,
            report_data_first_row=constants.REPORT_DATA_FIRST_ROW,
            total_work_duration_label=constants.TOTAL_WORK_DURATION_LABEL,
            col_section=constants.COL_SECTION,
            header_period_pattern=constants.REPORT_HEADER_PERIOD_PATTERN,
            mode_duration_columns=tuple(constants.MODE_DURATION_COLUMNS),
        )


@dataclass
class ModeDurationReportSectionRow:
    """Строка участка с длительностями режимов."""

    row_index: int
    section_name: str
    cells: Dict[str, str] = field(default_factory=dict)
    mode_duration_columns: Tuple[str, ...] = ()

    @property
    def mode_durations_seconds(self) -> Dict[str, int]:
        return {
            column_name: parse_duration_seconds(self.cells.get(column_name)) or 0
            for column_name in self.mode_duration_columns
        }

    @property
    def modes_sum_seconds(self) -> int:
        return sum(self.mode_durations_seconds.values())


@dataclass
class ModeDurationReportParsed:
    """Разобранный отчёт с таблицей участков и длительностями режимов."""

    title_info: ReportTitleInfo
    column_headers: List[str]
    section_rows: List[ModeDurationReportSectionRow]
    total_duration_seconds: Optional[int] = None
    total_duration_raw: str = ""
    total_label_row_index: Optional[int] = None


def parse_duration_seconds(value: object) -> Optional[int]:
    """Парсит длительность из ячейки (H:MM:SS, MM:SS или time/timedelta из Excel)."""
    if value is None:
        return None
    if isinstance(value, timedelta):
        return int(value.total_seconds())
    if isinstance(value, time):
        return (
            value.hour * TestConst.SECONDS_PER_HOUR
            + value.minute * TestConst.SEC_PER_MIN
            + value.second
        )
    if isinstance(value, datetime):
        return (
            value.hour * TestConst.SECONDS_PER_HOUR
            + value.minute * TestConst.SEC_PER_MIN
            + value.second
        )

    duration_text = _stringify_cell(value).strip()
    if not duration_text:
        return None

    parts = duration_text.split(":")
    try:
        if len(parts) == _DURATION_PARTS_H_MM_SS:
            hours, minutes, seconds = (int(part) for part in parts)
            return (
                hours * TestConst.SECONDS_PER_HOUR
                + minutes * TestConst.SEC_PER_MIN
                + seconds
            )
        if len(parts) == _DURATION_PARTS_MM_SS:
            minutes, seconds = (int(part) for part in parts)
            return minutes * TestConst.SEC_PER_MIN + seconds
    except ValueError:
        return None
    return None


def is_duration_cell_filled(value: object) -> bool:
    """Ячейка с длительностью заполнена (допускается 0:00:00)."""
    return parse_duration_seconds(value) is not None


def format_duration_seconds(total_seconds: int) -> str:
    """Форматирует длительность в секундах в строку H:MM:SS."""
    hours, remainder = divmod(total_seconds, TestConst.SECONDS_PER_HOUR)
    minutes, seconds = divmod(remainder, TestConst.SEC_PER_MIN)
    return f"{hours}:{minutes:02d}:{seconds:02d}"


def find_total_work_duration(
    worksheet: Worksheet,
    *,
    data_first_row: int,
    total_work_duration_label: str,
) -> Tuple[Optional[int], str, Optional[int]]:
    """
    Ищет строку с меткой суммарного времени и парсит длительность рядом.

    Возвращает (секунды, сырое значение ячейки, номер строки с меткой) или (None, "", None).
    """
    for row_index, row_values in enumerate(
        worksheet.iter_rows(min_row=data_first_row, values_only=True),
        start=data_first_row,
    ):
        for column_index, cell_value in enumerate(row_values):
            if cell_value is None:
                continue
            cell_text = _stringify_cell(cell_value).strip()
            if total_work_duration_label not in cell_text:
                continue

            duration_candidates = []
            if column_index + 1 < len(row_values):
                duration_candidates.append(row_values[column_index + 1])
            if row_index + 1 <= worksheet.max_row:
                duration_candidates.append(
                    worksheet.cell(row=row_index + 1, column=column_index + 1).value
                )

            for candidate in duration_candidates:
                duration_seconds = parse_duration_seconds(candidate)
                if duration_seconds is not None:
                    return duration_seconds, _stringify_cell(candidate).strip(), row_index

            return None, "", row_index

    return None, "", None


def parse_mode_duration_report_worksheet(
    worksheet: Worksheet,
    expected_section_names: List[str],
    layout: ModeDurationReportLayout,
) -> ModeDurationReportParsed:
    """
    Разбирает лист xlsx: шапка, колонки, строки участков и суммарное время.

    В section_rows попадают только участки из expected_section_names (без учёта регистра).
    """
    headers = get_report_column_headers(worksheet, layout.report_column_headers_row)
    title_info = parse_report_title(
        worksheet.cell(row=layout.report_title_row, column=1).value,
        layout.header_period_pattern,
    )
    total_duration_seconds, total_duration_raw, total_label_row_index = find_total_work_duration(
        worksheet,
        data_first_row=layout.report_data_first_row,
        total_work_duration_label=layout.total_work_duration_label,
    )

    section_rows: List[ModeDurationReportSectionRow] = []
    expected_names_lower = {name.lower() for name in expected_section_names}

    for row_index, row_values in enumerate(
        worksheet.iter_rows(
            min_row=layout.report_data_first_row,
            max_col=len(headers) if headers else 5,
            values_only=True,
        ),
        start=layout.report_data_first_row,
    ):
        if total_label_row_index is not None and row_index >= total_label_row_index:
            break

        cells = build_column_cells(row_values, headers)
        section_name = cells.get(layout.col_section, "").strip()
        if not section_name:
            continue
        if section_name.lower() not in expected_names_lower:
            continue

        section_rows.append(
            ModeDurationReportSectionRow(
                row_index=row_index,
                section_name=section_name,
                cells=cells,
                mode_duration_columns=layout.mode_duration_columns,
            )
        )

    return ModeDurationReportParsed(
        title_info=title_info,
        column_headers=headers,
        section_rows=section_rows,
        total_duration_seconds=total_duration_seconds,
        total_duration_raw=total_duration_raw,
        total_label_row_index=total_label_row_index,
    )


def format_mode_duration_section_rows_for_allure(
    section_rows: List[ModeDurationReportSectionRow],
) -> str:
    """Форматирует строки участков для вложения в Allure."""
    lines = []
    for row in section_rows:
        durations_text = ", ".join(
            f"{column}={format_duration_seconds(seconds)}"
            for column, seconds in row.mode_durations_seconds.items()
        )
        lines.append(
            f"row#{row.row_index}: {row.section_name} | sum={format_duration_seconds(row.modes_sum_seconds)} | "
            f"{durations_text}"
        )
    return "\n".join(lines)
