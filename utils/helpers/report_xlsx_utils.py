"""
Утилиты для разбора xlsx-отчётов и проверки их формата.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import allure
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from constants.test_constants import BaseTN3Constants as TestConst
from constants.test_constants import ExportReportConstants as ReportConst
from utils.helpers.ws_test_utils import extract_first_number, localize_as_moscow


@dataclass
class ReportTitleInfo:
    """Разобранная шапка отчёта"""

    raw_title: str
    period_start: Optional[datetime] = None
    period_end: Optional[datetime] = None


@dataclass
class LeakReportRow:
    """Разобранная строка данных по утечке"""

    row_index: int
    cells: Dict[str, str] = field(default_factory=dict)

    @property
    def datetime_value(self) -> Optional[datetime]:
        return parse_report_datetime(self.cells.get(ReportConst.COL_DATETIME))

    @property
    def object_value(self) -> str:
        return self.cells.get(ReportConst.COL_OBJECT, "")

    @property
    def lds_status(self) -> str:
        return self.cells.get(ReportConst.COL_LDS_STATUS, "")

    @property
    def masking_info(self) -> str:
        return self.cells.get(ReportConst.COL_MASK_INFO, "")

    @property
    def coordinate_meters(self) -> Optional[float]:
        coordinate_km = extract_first_number(self.cells.get(ReportConst.COL_COORDINATE))
        if coordinate_km is None:
            return None
        return coordinate_km * TestConst.KM_TO_METERS

    @property
    def leak_volume(self) -> Optional[float]:
        return extract_first_number(self.cells.get(ReportConst.COL_LEAK_VOLUME))

    @property
    def mt_mode(self) -> str:
        return self.cells.get(ReportConst.COL_MT_MODE, "")


def is_xlsx_file_bytes(file_bytes: Optional[bytes]) -> bool:
    """Проверяет zip-сигнатуру xlsx"""
    if not file_bytes:
        return False
    return file_bytes.startswith(ReportConst.ZIP_SIGNATURE)


def is_xlsx_extension(file_name: str) -> bool:
    """Проверяет расширение .xlsx без учёта регистра."""
    return file_name.lower().endswith(ReportConst.XLSX_EXTENSION)


def parse_report_datetime(value: object) -> Optional[datetime]:
    """Парсит дату/время из ячейки отчёта."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), ReportConst.REPORT_DATETIME_FORMAT)
        except ValueError:
            return None
    return None


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(ReportConst.REPORT_DATETIME_FORMAT)
    return str(value)


def normalize_report_period_naive(value: datetime) -> datetime:
    """Московское время без tzinfo и микросекунд - для сравнения периодов в отчёте."""
    return localize_as_moscow(value).replace(microsecond=0, tzinfo=None)


def report_period_comparison_bounds(
    period_start: datetime,
    period_end: datetime,
    tolerance_minutes: int = ReportConst.REPORT_PERIOD_TOLERANCE_MINUTES,
) -> tuple[datetime, datetime, datetime, datetime]:
    """
    Границы периода с допуском +-tolerance_minutes для start и end отдельно.
    Возвращает (start_lower, start_upper, end_lower, end_upper).
    """
    start = normalize_report_period_naive(period_start)
    end = normalize_report_period_naive(period_end)
    delta = timedelta(minutes=tolerance_minutes)
    return start - delta, start + delta, end - delta, end + delta


def build_export_report_file_name(
    tu_description: str,
    period_start: datetime,
    period_end: datetime,
    report_name_part: str = ReportConst.LEAKS_REPORT_NAME_PART,
    name_tu_separator: str = " ",
) -> str:
    """
    Имя xlsx при скачивании: '{название}{sep}{ТУ} DD.MM.YYYY HH_MM_SS - DD.MM.YYYY HH_MM_SS.xlsx'.
    По умолчанию - отчёт об утечках.
    """
    start_text = normalize_report_period_naive(period_start).strftime(ReportConst.REPORT_FILE_NAME_DATETIME_FORMAT)
    end_text = normalize_report_period_naive(period_end).strftime(ReportConst.REPORT_FILE_NAME_DATETIME_FORMAT)
    return (
        f"{report_name_part}{name_tu_separator}{tu_description} {start_text} - {end_text}"
        f"{ReportConst.XLSX_EXTENSION}"
    )


def parse_period_from_export_file_name(
    file_name: str,
    period_pattern: str | None = None,
) -> tuple[Optional[datetime], Optional[datetime]]:
    """Извлекает границы периода из имени скачанного xlsx-файла."""
    match = re.search(
        period_pattern or ReportConst.REPORT_FILE_NAME_PERIOD_PATTERN,
        file_name.strip(),
        re.IGNORECASE,
    )
    if match is None:
        return None, None

    parse_format = ReportConst.REPORT_FILE_NAME_DATETIME_FORMAT.replace("_", ":")

    def _parse_part(value: str) -> Optional[datetime]:
        try:
            return datetime.strptime(value.replace("_", ":"), parse_format)
        except ValueError:
            return None

    return _parse_part(match.group("period_start")), _parse_part(match.group("period_end"))


def parse_report_title(
    title_raw: object,
    header_period_pattern: str | None = None,
) -> ReportTitleInfo:
    """
    Парсит шапку отчёта с именованными группами period_start/period_end.
    """
    title_str = _stringify_cell(title_raw)
    pattern = header_period_pattern or ReportConst.REPORT_HEADER_PERIOD_PATTERN
    match = re.search(pattern, title_str)
    if match is None:
        return ReportTitleInfo(raw_title=title_str)

    return ReportTitleInfo(
        raw_title=title_str,
        period_start=parse_report_datetime(match.group("period_start")),
        period_end=parse_report_datetime(match.group("period_end")),
    )


def load_report_worksheet(file_path: Path) -> Optional[Worksheet]:
    """Открывает первый лист xlsx. При ошибке возвращает None."""
    if not file_path.exists():
        return None
    try:
        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    except Exception:
        return None
    sheet_names = workbook.sheetnames
    if not sheet_names:
        return None
    return workbook[sheet_names[ReportConst.DEFAULT_SHEET_INDEX]]


def get_report_title_cell(worksheet: Worksheet) -> object:
    return worksheet.cell(row=ReportConst.REPORT_TITLE_ROW, column=1).value


def get_report_column_headers(
    worksheet: Worksheet,
    headers_row: int = ReportConst.REPORT_COLUMN_HEADERS_ROW,
) -> List[str]:
    """Возвращает непустые заголовки колонок из указанной строки шапки."""
    headers: List[str] = []
    column_index = 1
    while True:
        cell_value = worksheet.cell(row=headers_row, column=column_index).value
        if cell_value is None or not str(cell_value).strip():
            break
        headers.append(_stringify_cell(cell_value).strip())
        column_index += 1
    return headers


def build_column_cells(row_values: tuple, headers: List[str]) -> Dict[str, str]:
    """Собирает словарь {название колонки: значение ячейки} по строке данных."""
    return {
        header: _stringify_cell(row_values[column_index]) if column_index < len(row_values) else ""
        for column_index, header in enumerate(headers)
    }


def iter_report_data_rows(worksheet: Worksheet) -> List[LeakReportRow]:
    """
    Возвращает строки данных по утечкам, начиная с REPORT_DATA_FIRST_ROW.
    Пустые строки пропускаются.
    """
    headers = get_report_column_headers(worksheet)
    if not headers:
        return []

    rows: List[LeakReportRow] = []
    for excel_row_index, row_values in enumerate(
        worksheet.iter_rows(
            min_row=ReportConst.REPORT_DATA_FIRST_ROW,
            max_col=len(headers),
            values_only=True,
        ),
        start=ReportConst.REPORT_DATA_FIRST_ROW,
    ):
        if not any(cell is not None and str(cell).strip() for cell in row_values):
            continue
        rows.append(
            LeakReportRow(
                row_index=excel_row_index,
                cells=build_column_cells(row_values, headers),
            )
        )
    return rows


def find_row_with_object(rows: List[LeakReportRow], object_substring: str) -> Optional[LeakReportRow]:
    """Ищет первую строку, где колонка 'Объект' содержит подстроку без учёта регистра"""
    substring_lower = object_substring.lower()
    for row in rows:
        if substring_lower in row.object_value.lower():
            return row
    return None


def read_worksheet_cell_value(
    file_path: Path,
    row: int,
    column: int,
    *,
    data_only: bool = True,
    sheet_index: int = ReportConst.DEFAULT_SHEET_INDEX,
) -> object:
    """
    Читает значение ячейки из xlsx.

    Для формул используйте data_only=False, иначе openpyxl вернёт вычисленное значение.
    """
    if not file_path.exists():
        return None
    try:
        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=data_only)
    except Exception:
        return None
    sheet_names = workbook.sheetnames
    if not sheet_names:
        return None
    worksheet = workbook[sheet_names[sheet_index]]
    return worksheet.cell(row=row, column=column).value


def read_worksheet_cell_formula(
    file_path: Path,
    row: int,
    column: int,
    *,
    sheet_index: int = ReportConst.DEFAULT_SHEET_INDEX,
) -> str:
    """
    Читает формулу ячейки из xlsx.

    read_only-режим openpyxl не возвращает формулы, поэтому загрузка выполняется
    с read_only=False и data_only=False.
    """
    if not file_path.exists():
        return ""
    workbook = None
    try:
        workbook = load_workbook(filename=str(file_path), read_only=False, data_only=False)
        sheet_names = workbook.sheetnames
        if not sheet_names:
            return ""
        worksheet = workbook[sheet_names[sheet_index]]
        value = worksheet.cell(row=row, column=column).value
        return value if isinstance(value, str) else ""
    except Exception:
        return ""
    finally:
        if workbook is not None:
            workbook.close()


def sum_duration_columns_across_rows(
    section_rows: list,
    mode_duration_columns: list[str],
) -> dict[str, int]:
    """Суммирует длительности по колонкам режимов для всех строк участков."""
    totals = {column_name: 0 for column_name in mode_duration_columns}
    for section_row in section_rows:
        for column_name, duration_seconds in section_row.mode_durations_seconds.items():
            totals[column_name] += duration_seconds
    return totals


def save_report_bytes_to_temp_file(
    file_bytes: bytes,
    prefix: str = "leaks_report_",
) -> Optional[Path]:
    """Сохраняет байты отчёта во временный xlsx-файл. При ошибке возвращает None."""
    import tempfile

    try:
        with tempfile.NamedTemporaryFile(
            suffix=ReportConst.XLSX_EXTENSION,
            prefix=prefix,
            delete=False,
        ) as temp_file:
            temp_file.write(file_bytes)
            return Path(temp_file.name)
    except OSError:
        return None


def attach_report_file_to_allure(file_path: Path, file_name: str) -> None:
    """Прикладывает xlsx к Allure при падении теста"""
    try:
        xlsx_type = allure.attachment_type.XLSX
    except AttributeError:
        xlsx_type = None

    if xlsx_type is not None:
        allure.attach.file(
            str(file_path),
            name=file_name,
            attachment_type=xlsx_type,
            extension="xlsx",
        )
        return

    try:
        with file_path.open("rb") as raw_file:
            allure.attach(raw_file.read(), name=file_name, extension="xlsx")
    except OSError:
        pass
